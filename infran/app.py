import time

from flask import Flask, render_template, request
import requests
from bs4 import BeautifulSoup


#엑셀을 쓰기위한 준비
from openpyxl import Workbook
write_wb = Workbook()
write_ws = write_wb.active

# write_ws.cell(1,1,"안녕")
# write_wb.save("result.xlsx")


#셀레늄 import
from selenium import webdriver

app = Flask(__name__)

@app.route('/')
def hello_world():
    return render_template("index.html")

@app.route('/result', methods=['POST'])
def result():
    if request.method == 'POST':

    #form 에서 받은값을input1,2를 크롤링해서 result.html로 넘겨줌줌
        keyword = request.form['input1']
        page = request.form['input2']

    #https://search.daum.net/search?w=tot&DA=YZR&t__nil_searchbox=btn&sug=&sugo=&q=
    # %EC%BD%94%EB%A1%9C%EB%82%9819

        daum_list =[]

        url = "https://search.daum.net/search?w=tot&DA=YZR&t__nil_searchbox=btn&sug=&sugo=&q="+keyword


        for i in range(1, int(page)+1):
            req = requests.get(url+"&p="+str(i))
            soup = BeautifulSoup(req.text, 'html.parser')
            for i in soup.find_all("a", class_="f_link_b"):
                print(i.text)
                daum_list.append(i.text)

        for i in range(1, len(daum_list)+1):
            write_ws.cell(i,1,daum_list[i-1])

        write_wb.save("static/result.xlsx")
        # print(request.form['input1'])
        # print(request.form['input2'])

        return render_template("result.html", daum_list=daum_list)



@app.route('/naver_shopping', methods=['POST'])
def naver_shopping():

    search_list =[]
    search_list_src =[]


    search = request.form['input3']
    driver = webdriver.Chrome()
    driver.implicitly_wait(3)
    driver.get("https://search.shopping.naver.com/search/all.nhn?query="+search+"&cat_id=&frm=NVSHATC")


#스크롤 내려서 사진 불러오기
    y = 1000
    for timer in range(1,10):
        driver.execute_script("window.scrollTo(0," + str(y)+ ")")
        y= y+1000
        time.sleep(1)


    soup = BeautifulSoup(driver.page_source,"html.parser")

    for i in soup.select("#_search_list > div.search_list.basis > ul > li"):
        print(i.find("a", class_="link").text)

        search_list.append(i.find("a", class_="link").text)
        search_list_src.append(i.find("img", class_="_productLazyImg")['src'])

    print("_0___핫딜   -0-0-0-0-0--0-0-0-0-0-0-0-")


    driver.find_element_by_class_name("snb_hotdeal").click()

    # 스크롤 내려서 사진 불러오기
    y = 1000
    for timer in range(1, 10):
        driver.execute_script("window.scrollTo(0," + str(y) + ")")
        y = y + 1000
        time.sleep(1)


    soup = BeautifulSoup(driver.page_source,"html.parser")

    for i in soup.select("#_search_list > div.search_list.basis > ul > li"):
        print(i.find("a",class_="link").text)
        search_list.append(i.find("a", class_="link").text)
        search_list_src.append(i.find("img", class_="_productLazyImg")['src'])

    print("----0-0-0-0-0-0-0-0-0-0-0-0-0-0-0-해외직구")

    driver.find_element_by_class_name("snb_overseas").click()

    # 스크롤 내려서 사진 불러오기
    y = 1000
    for timer in range(1, 10):
        driver.execute_script("window.scrollTo(0," + str(y) + ")")
        y = y + 1000
        time.sleep(1)


    soup = BeautifulSoup(driver.page_source, "html.parser")

    for i in soup.select("#_search_list > div.search_list.basis > ul > li"):
        print(i.find("a", class_="link").text)
        search_list.append(i.find("a", class_="link").text)
        search_list_src.append(i.find("img", class_="_productLazyImg")['src'])
    #driver.close()

    return render_template("shopping.html", search_list=search_list,
                           search_list_src=search_list_src,
                           len= len(search_list_src))

if __name__ == '__main__':
    app.debug = True
    app.run(host='0.0.0.0')
    # app.run(debug=True)


